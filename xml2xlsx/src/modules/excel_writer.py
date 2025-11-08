"""Excel writer."""

from typing import cast
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

from .xml_reader import XMLReader
from .config import ConfigData


def is_number(s: str) -> bool:
    """Check if the string is castable to float or not."""
    try:
        float(s)
    except ValueError:
        return False
    return True


class ExcelWriter:
    """Write data to an Excel file.

    Methods: __init__: Constructor.
             _reset_row: Clear row for the next build row call.
             _build_row: Construct the row to write from the XML file.
             columns_are_correct: Check if the Excel columns are as expected.
             _get_rr_interval: Convert RRinterval to the correct format.
             _update_cell: Update one cell in the row with the right formatting.
             _update_row: Update a row in the Excel file with data from the XML file.
    """

    RR_INTERVAl_COLUMN: int = 7
    FLOAT_COLUMNS: set[int] = {8, 9, 10, 11, 12, 17, 18, 19}
    INTEGER_COLUMNS: set[int] = {21, 22, 23, 24, 31}

    def __init__(self, config_data: ConfigData) -> None:
        """Initialize the object.

        :param: config_data: Dictionary with configuration data.
        :ivar:  _file_path: MS Excel file to read.
                _workbook: Excel workbook object.
                _sheet: Sheet (tab) to work on in the Excel file.
                _row: Row to write to the Excel file.
                _row_number: Number of the row to write to the Excel file.
                _mapping: Excel column number mapped to XLM tags.
                _excel_headings: Excel column headings.
        """
        self._file_path: str = cast(str, config_data["excel_path"])
        # pylint: disable=too-many-try-statements
        try:
            self._workbook = load_workbook(self._file_path)
            self._sheet = self._workbook[cast(str, config_data["sheet_title"])]
        except FileNotFoundError as e:
            print(f"ERROR: File {self._file_path} not found.")
            raise SystemExit from e
        except KeyError as e:
            print(f'ERROR: Sheet {config_data["sheet_title"]} not found in {self._file_path}.')
            raise SystemExit from e

        self._row: list[str] = [""] * 32
        self._row_number: int = 0
        # pylint: disable=too-many-try-statements
        try:
            # Maps Excel column number to XML tags
            self._mapping: dict[int, str] = cast(dict[int, str], config_data["mapping"])
            self._excel_headings: list[str] = cast(list[str], config_data["excel_headings"])
        except KeyError as e:
            print("ERROR: Missing key (mapping or excel_headings) in config.yaml. Stopping without any action taken.")
            raise SystemExit from e

    def _reset_row(self) -> None:
        """Reset row to write, after writing."""
        self._row = [""] * 32
        self._row_number = 0

    def _build_row(self, xml_file_path: str) -> None:
        """Build a list of data to add to an Excel row.

        :param: xml_file_path: XML file for which to build the row in Excel.
        """
        xml_reader: XMLReader = XMLReader(xml_file_path, list(self._mapping.values()))
        xml_heading_to_value: dict[str, str] = xml_reader.get_values()
        for col_number, xml_heading in self._mapping.items():
            self._row[col_number - 1] = xml_heading_to_value[xml_heading]  # Excel columns start at 1, row list at 0.
        # Set row number
        for cell in self._sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=False):
            if cell[0].value == self._row[1]:  # Patient ID.
                # As check in the if statement assures .row is an integer, it is safe to cast.
                self._row_number = cell[0].row

    def columns_are_correct(self) -> bool:
        """Check columns in Excel are as expected.

        Avoids wrong writes if a column was added or removed.
        :return: True if the columns are as expected, False otherwise.
        """
        # pylint: disable=consider-using-any-or-all
        for index, column in enumerate(self._mapping):
            if self._sheet.cell(row=1, column=column).value != self._excel_headings[index]:
                return False
        return True

    def _get_rr_interval(self) -> str:
        """Put RRinterval in the right format.

        :return: RRinterval as a string in the wanted format.
        """
        # Convert from the format '89,0%' to '89.0'
        self._row[6] = self._row[6].replace("%", "")
        self._row[6] = self._row[6].replace(",", ".")
        if is_number(self._row[6]):
            rr_interval_value: float = float(self._row[6]) / 100
            if 0 <= rr_interval_value <= 1:  # Can be written as %
                return str(rr_interval_value)
        print(f"WARNING: UNPLAUSIBLE RRInterval {self._row[6]} for Patient ID {self._row[1]}.")
        return "UNPLAUSIBLE"

    def _update_cell(self, column: int, value: str, formatting: str) -> None:
        """Update one cell in the row with the right formatting.

        Formatting documentation:
        https://pythoninoffice.com/python-excel-number-format/
        :param: column: Column number to write to.
                value: Value to write to the cell.
                formatting: Formatting to apply to the cell.
        """
        color: str = "FFFF00"  # Yellow fill, TODO: make configurable.
        if is_number(value):
            value = float(value)  # type: ignore[assignment] # Intended here.
        self._sheet.cell(column=column, row=self._row_number, value=value)
        self._sheet.cell(column=column, row=self._row_number).number_format = formatting
        self._sheet.cell(column=column, row=self._row_number).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )

    # pylint: disable=too-complex
    def update_row(self, xml_file_path: str) -> None:
        """Update a row in the Excel file with data from the XML file.

        :param: xml_file_path: XML path for which to update the row in Excel.
        """
        self._build_row(xml_file_path)
        if self._row_number not in [0, 1]:  # Update the row.
            for index, value in enumerate(self._row):
                column = index + 1  # Excel starts counting at 1
                # Only write if the cell is empty and there is a value to write.
                if not self._sheet.cell(column=column, row=self._row_number).value and value != "":
                    format_parameter: str = "General"  # For strings
                    if column == ExcelWriter.RR_INTERVAl_COLUMN:
                        # Overwrite the value. Then it can be formatted correctly.
                        value = self._get_rr_interval()  # pylint: disable=redefined-loop-name.
                    if is_number(value):
                        match column:
                            case ExcelWriter.RR_INTERVAl_COLUMN:  # Percentage.
                                format_parameter = "0%"
                            case col if col in ExcelWriter.FLOAT_COLUMNS:  # Floats with 1 decimal.
                                format_parameter = "0.0"
                            case col if col in ExcelWriter.INTEGER_COLUMNS:  # Integers.
                                format_parameter = "0"
                    self._update_cell(column, value, format_parameter)
            try:
                self._workbook.save(self._file_path)
            except PermissionError as exc:
                print(f"ERROR: Permission denied for {self._file_path}. Is the Excel file open? Please close it.")
                raise SystemExit from exc
        else:  # The Patient ID was not found in the Excel file.
            print(
                f"WARNING: Patient ID {self._row[1]} from XML file {xml_file_path} was not found in "
                + f"Excel file {self._file_path}."
            )
        self._reset_row()


if __name__ == "__main__":
    pass
